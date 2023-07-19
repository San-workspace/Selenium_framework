import openpyxl
class HomepageData:
    Test_Home_Data = [{"name":"san1", "email":"1gmail.com", "dob":"1988"}, {"name":"alwar", "email":"ymail.com","dob":"1992"}]

    @staticmethod
    def gettestdata(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\DEEPTHA\\Desktop\\Sankar\\Selenium-Python\\Testdata.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):  # iterating rows
            if (sheet.cell(row=i, column=1).value) == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[dict]    #pytest will accept the data in list form so converting dic to list