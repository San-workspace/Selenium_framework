import openpyxl
book=openpyxl.load_workbook("C:\\Users\\DEEPTHA\\Desktop\\Sankar\\Selenium-Python\\Testdata.xlsx")
sheet=book.active
# cell=sheet.cell(row=2,column=1)
# print(cell.value)

#write data to specified cell
# # sheet.cell(row=1,column=3).value="san"
# # print(sheet.cell(row=2,column=3).value)

sheet['B10'].value="san-msys"
print(sheet['B10'].value)
book.save("C:\\Users\\DEEPTHA\\Desktop\\Sankar\\Selenium-Python\\Testdata.xlsx")


#others :
# print(sheet.max_row)                              #to get max rows
# print(sheet.max_column)                           #to get max col
# print(sheet['C1'].value)                          #to get specified cell val

#creating dictonary as a key:value pair from xl data :

# dict={}
# for i in range(1,sheet.max_row+1):       #iterating rows
#     if(sheet.cell(row=i,column=1).value)=="test2":  #if wants to get only one testcase data
#         for j in range(2,sheet.max_column+1):           #getting columns values in each test (from col 2)
#             dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value  #generating dic by using get values
# print(dict)

